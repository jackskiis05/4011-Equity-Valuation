import warnings
from dataclasses import dataclass

import numpy as np
import pandas as pd
import plotly.express as px
import streamlit as st

warnings.filterwarnings("ignore")

# -----------------------------
# Runtime guard
# -----------------------------
def _running_under_streamlit() -> bool:
    try:
        from streamlit.runtime import exists as _rt_exists
        return _rt_exists()
    except Exception:
        try:
            from streamlit.runtime.scriptrunner import get_script_run_ctx
            return get_script_run_ctx() is not None
        except Exception:
            return False


if not _running_under_streamlit():
    import subprocess
    import sys
    script_path = sys.argv[0] if sys.argv and sys.argv[0] else __file__
    print(
        "\n" + "=" * 70 + "\n"
        "This is a Streamlit app. Launching it via the Streamlit CLI now...\n"
        f"    streamlit run \"{script_path}\"\n"
        "The app will open in your browser at http://localhost:8501\n"
        "Press Ctrl+C in the terminal to stop it.\n"
        + "=" * 70 + "\n",
        flush=True,
    )
    try:
        subprocess.run(
            [sys.executable, "-m", "streamlit", "run", script_path],
            check=False,
        )
    except FileNotFoundError:
        print(
            "Could not launch Streamlit. Install it with:\n"
            "    pip install streamlit yfinance plotly pandas numpy openpyxl\n"
            "Then run manually:\n"
            f"    streamlit run \"{script_path}\""
        )
    sys.exit(0)

try:
    import yfinance as yf
except Exception:
    yf = None

st.set_page_config(page_title="Equity Valuation App", layout="wide")


# =========================================================
# Helpers
# =========================================================
@dataclass
class DCFInputs:
    ticker: str
    company_name: str
    revenue_0: float
    growth_years_1_5: list
    ebit_margin_0: float
    ebit_margin_target: float
    tax_rate: float
    da_pct_revenue: float
    capex_pct_revenue: float
    nwc_pct_revenue: float
    wacc: float
    terminal_growth: float
    shares_outstanding: float
    net_debt: float
    current_price: float


def pct_from_ui(x: float) -> float:
    return x / 100.0


def fmt_currency(x: float) -> str:
    if x is None or pd.isna(x):
        return "N/A"
    return f"${x:,.2f}"


def fmt_number(x: float) -> str:
    if x is None or pd.isna(x):
        return "N/A"
    return f"{x:,.2f}"


def fmt_pct(x: float) -> str:
    if x is None or pd.isna(x):
        return "N/A"
    return f"{x * 100:.2f}%"


def safe_float(value, default=np.nan):
    try:
        if value is None:
            return default
        return float(value)
    except Exception:
        return default


def _first_available(df: pd.DataFrame, candidates: list):
    if not isinstance(df, pd.DataFrame) or df.empty:
        return np.nan
    for idx in candidates:
        if idx in df.index:
            val = safe_float(df.loc[idx].iloc[0])
            if not pd.isna(val):
                return val
    return np.nan


@st.cache_data(show_spinner=False)
def get_stock_data(ticker: str):
    if yf is None:
        return {
            "success": False,
            "error": "yfinance is not installed. Run: pip install yfinance",
        }

    diagnostics = []
    try:
        stock = yf.Ticker(ticker)

        try:
            info = stock.info or {}
            if not info:
                diagnostics.append("stock.info returned empty")
        except Exception as e:
            info = {}
            diagnostics.append(f"stock.info failed: {e}")

        financials = pd.DataFrame()
        try:
            financials = stock.income_stmt
            if not isinstance(financials, pd.DataFrame) or financials.empty:
                diagnostics.append("stock.income_stmt empty, trying financials")
                financials = stock.financials
                if not isinstance(financials, pd.DataFrame) or financials.empty:
                    diagnostics.append("stock.financials also empty")
        except Exception as e:
            diagnostics.append(f"income statement fetch failed: {e}")
            try:
                financials = stock.financials
            except Exception as e2:
                diagnostics.append(f"financials fallback failed: {e2}")
                financials = pd.DataFrame()

        try:
            balance_sheet = stock.balance_sheet
            if not isinstance(balance_sheet, pd.DataFrame) or balance_sheet.empty:
                diagnostics.append("balance_sheet is empty")
        except Exception as e:
            diagnostics.append(f"balance_sheet fetch failed: {e}")
            balance_sheet = pd.DataFrame()

        # Cash flow statement — used for historical capex and D&A.
        try:
            cashflow = stock.cashflow
            if not isinstance(cashflow, pd.DataFrame) or cashflow.empty:
                diagnostics.append("cashflow statement is empty")
        except Exception as e:
            diagnostics.append(f"cashflow fetch failed: {e}")
            cashflow = pd.DataFrame()

        # Sector/industry — useful context for users.
        sector = info.get("sector") or "N/A"
        industry = info.get("industry") or "N/A"

        company_name = info.get("shortName") or info.get("longName") or ticker.upper()
        current_price = safe_float(info.get("currentPrice"))
        market_cap = safe_float(info.get("marketCap"))
        shares = safe_float(info.get("sharesOutstanding"))

        revenue = _first_available(financials, ["Total Revenue", "Operating Revenue"])
        ebit = _first_available(financials, ["EBIT", "Operating Income"])

        total_debt = _first_available(
            balance_sheet,
            ["Total Debt", "Current Debt And Capital Lease Obligation", "Long Term Debt"],
        )
        cash = _first_available(
            balance_sheet,
            [
                "Cash And Cash Equivalents",
                "Cash Cash Equivalents And Short Term Investments",
                "Cash",
            ],
        )
        current_assets = _first_available(balance_sheet, ["Current Assets"])
        current_liabilities = _first_available(balance_sheet, ["Current Liabilities"])

        nwc_estimate = (
            current_assets - current_liabilities
            if not pd.isna(current_assets) and not pd.isna(current_liabilities)
            else np.nan
        )

        net_debt = (
            total_debt - cash
            if not pd.isna(total_debt) and not pd.isna(cash)
            else np.nan
        )
        ebit_margin = (
            ebit / revenue
            if not pd.isna(revenue) and revenue != 0 and not pd.isna(ebit)
            else np.nan
        )
        nwc_pct_revenue = (
            nwc_estimate / revenue
            if not pd.isna(revenue) and revenue != 0 and not pd.isna(nwc_estimate)
            else np.nan
        )

        try:
            history = stock.history(period="1y")
        except Exception:
            history = pd.DataFrame()

        # ----- Historical metrics across all available years -----
        # yfinance income statements have columns = fiscal year-end dates (most recent first).
        def _row_series(df: pd.DataFrame, candidates: list) -> pd.Series:
            """Return a numeric series for the first matching row, oldest -> newest."""
            if not isinstance(df, pd.DataFrame) or df.empty:
                return pd.Series(dtype=float)
            for idx in candidates:
                if idx in df.index:
                    s = pd.to_numeric(df.loc[idx], errors="coerce").dropna()
                    if not s.empty:
                        # Sort by date (column labels) so oldest is first
                        try:
                            s = s.sort_index()
                        except Exception:
                            pass
                        return s
            return pd.Series(dtype=float)

        rev_series = _row_series(financials, ["Total Revenue", "Operating Revenue"])
        ebit_series = _row_series(financials, ["EBIT", "Operating Income"])
        capex_series = _row_series(cashflow, ["Capital Expenditure", "Capital Expenditures"])
        da_series = _row_series(
            cashflow,
            ["Depreciation And Amortization", "Depreciation Amortization Depletion",
             "Reconciled Depreciation", "Depreciation"],
        )

        # Historical revenue growth: year-over-year % change, average across available years.
        hist_rev_growth = np.nan
        hist_rev_growth_list = []
        if len(rev_series) >= 2:
            growths = rev_series.pct_change().dropna()
            if not growths.empty:
                hist_rev_growth = float(growths.mean())
                hist_rev_growth_list = [float(x) for x in growths.tolist()]

        # Historical EBIT margin: mean/min/max of ebit/revenue per year.
        hist_ebit_margins = []
        if not rev_series.empty and not ebit_series.empty:
            common = rev_series.index.intersection(ebit_series.index)
            for col in common:
                rv = rev_series.get(col)
                eb = ebit_series.get(col)
                if rv and rv != 0 and eb is not None and not pd.isna(eb):
                    hist_ebit_margins.append(eb / rv)

        hist_ebit_margin_mean = float(np.mean(hist_ebit_margins)) if hist_ebit_margins else np.nan
        hist_ebit_margin_min = float(np.min(hist_ebit_margins)) if hist_ebit_margins else np.nan
        hist_ebit_margin_max = float(np.max(hist_ebit_margins)) if hist_ebit_margins else np.nan

        # Historical D&A and capex as % of revenue.
        hist_da_pct = np.nan
        hist_capex_pct = np.nan
        if not rev_series.empty:
            if not da_series.empty:
                common_da = rev_series.index.intersection(da_series.index)
                ratios = []
                for col in common_da:
                    rv = rev_series.get(col)
                    if rv and rv != 0:
                        ratios.append(abs(da_series.get(col, 0)) / rv)
                if ratios:
                    hist_da_pct = float(np.mean(ratios))
            if not capex_series.empty:
                common_cx = rev_series.index.intersection(capex_series.index)
                ratios = []
                for col in common_cx:
                    rv = rev_series.get(col)
                    if rv and rv != 0:
                        # Capex is reported negative on the cash flow statement; take abs
                        ratios.append(abs(capex_series.get(col, 0)) / rv)
                if ratios:
                    hist_capex_pct = float(np.mean(ratios))

        return {
            "success": True,
            "ticker": ticker.upper(),
            "company_name": company_name,
            "sector": sector,
            "industry": industry,
            "current_price": current_price,
            "market_cap": market_cap,
            "shares_outstanding": shares,
            "revenue": revenue,
            "ebit": ebit,
            "ebit_margin": ebit_margin,
            "cash": cash,
            "total_debt": total_debt,
            "net_debt": net_debt,
            "nwc_pct_revenue": nwc_pct_revenue,
            "price_history": history,
            "diagnostics": diagnostics,
            # Historical benchmarks
            "hist_revenue_growth_avg": hist_rev_growth,
            "hist_revenue_growth_list": hist_rev_growth_list,
            "hist_ebit_margin_mean": hist_ebit_margin_mean,
            "hist_ebit_margin_min": hist_ebit_margin_min,
            "hist_ebit_margin_max": hist_ebit_margin_max,
            "hist_da_pct": hist_da_pct,
            "hist_capex_pct": hist_capex_pct,
        }

    except Exception as e:
        return {
            "success": False,
            "error": f"Could not retrieve data for {ticker.upper()}: {e}",
            "diagnostics": diagnostics,
        }


def build_forecast(inputs: DCFInputs, years: int = 5) -> pd.DataFrame:
    rows = []
    revenue_prev = inputs.revenue_0
    nwc_prev = revenue_prev * inputs.nwc_pct_revenue

    for year in range(1, years + 1):
        growth = inputs.growth_years_1_5[year - 1]
        revenue = revenue_prev * (1 + growth)
        ebit_margin = inputs.ebit_margin_0 + (
            inputs.ebit_margin_target - inputs.ebit_margin_0
        ) * (year / years)
        ebit = revenue * ebit_margin
        nopat = ebit * (1 - inputs.tax_rate)
        da = revenue * inputs.da_pct_revenue
        capex = revenue * inputs.capex_pct_revenue
        nwc = revenue * inputs.nwc_pct_revenue
        delta_nwc = nwc - nwc_prev
        fcff = nopat + da - capex - delta_nwc

        rows.append({
            "Year": year, "Revenue": revenue, "Revenue Growth": growth,
            "EBIT Margin": ebit_margin, "EBIT": ebit, "NOPAT": nopat,
            "D&A": da, "Capex": capex, "NWC": nwc,
            "Change in NWC": delta_nwc, "FCFF": fcff,
        })
        revenue_prev = revenue
        nwc_prev = nwc

    return pd.DataFrame(rows)


def discount_cash_flows(df: pd.DataFrame, wacc: float, terminal_growth: float):
    out = df.copy()
    out["Discount Factor"] = 1 / ((1 + wacc) ** out["Year"])
    out["PV of FCFF"] = out["FCFF"] * out["Discount Factor"]

    final_fcff = out.iloc[-1]["FCFF"]
    terminal_value = final_fcff * (1 + terminal_growth) / (wacc - terminal_growth)
    pv_terminal_value = terminal_value / ((1 + wacc) ** out.iloc[-1]["Year"])
    pv_fcf = out["PV of FCFF"].sum()
    enterprise_value = pv_fcf + pv_terminal_value

    return out, terminal_value, pv_terminal_value, pv_fcf, enterprise_value


def run_dcf(inputs: DCFInputs):
    forecast = build_forecast(inputs)
    discounted, terminal_value, pv_terminal_value, pv_fcf, enterprise_value = discount_cash_flows(
        forecast, inputs.wacc, inputs.terminal_growth
    )
    equity_value = enterprise_value - inputs.net_debt
    value_per_share = (
        equity_value / inputs.shares_outstanding
        if inputs.shares_outstanding and inputs.shares_outstanding > 0
        else np.nan
    )
    premium_discount = (
        (value_per_share / inputs.current_price - 1)
        if (
            not pd.isna(inputs.current_price)
            and inputs.current_price > 0
            and not pd.isna(value_per_share)
        )
        else np.nan
    )
    return {
        "forecast": forecast, "discounted": discounted,
        "terminal_value": terminal_value, "pv_terminal_value": pv_terminal_value,
        "pv_fcf": pv_fcf, "enterprise_value": enterprise_value,
        "equity_value": equity_value, "value_per_share": value_per_share,
        "premium_discount": premium_discount,
    }


def make_sensitivity_table(inputs: DCFInputs):
    wacc_range = np.round(np.arange(max(0.01, inputs.wacc - 0.02), inputs.wacc + 0.021, 0.01), 4)
    tg_range = np.round(
        np.arange(max(0.0, inputs.terminal_growth - 0.01), inputs.terminal_growth + 0.011, 0.005),
        4,
    )
    table = pd.DataFrame(index=tg_range, columns=wacc_range, dtype=float)
    for tg in tg_range:
        for w in wacc_range:
            if tg >= w:
                table.loc[tg, w] = np.nan
            else:
                temp_inputs = DCFInputs(
                    ticker=inputs.ticker, company_name=inputs.company_name,
                    revenue_0=inputs.revenue_0, growth_years_1_5=inputs.growth_years_1_5,
                    ebit_margin_0=inputs.ebit_margin_0, ebit_margin_target=inputs.ebit_margin_target,
                    tax_rate=inputs.tax_rate, da_pct_revenue=inputs.da_pct_revenue,
                    capex_pct_revenue=inputs.capex_pct_revenue, nwc_pct_revenue=inputs.nwc_pct_revenue,
                    wacc=float(w), terminal_growth=float(tg),
                    shares_outstanding=inputs.shares_outstanding, net_debt=inputs.net_debt,
                    current_price=inputs.current_price,
                )
                table.loc[tg, w] = round(run_dcf(temp_inputs)["value_per_share"], 2)
    table.index.name = "Terminal Growth"
    table.columns.name = "WACC"
    return table


def build_excel_bytes(inputs: DCFInputs, results: dict, sensitivity_df: pd.DataFrame) -> bytes:
    """Package the valuation into a multi-sheet Excel workbook with LIVE FORMULAS.

    The Forecast, DCF Bridge, and Summary sheets reference the Assumptions
    sheet so the user can edit any assumption in Excel and watch every
    downstream number recalculate automatically.

    The Sensitivity sheet is written as static values because it's a what-if
    grid by definition.
    """
    from io import BytesIO

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl is required to write formula-based Excel files. "
            "Install with: pip install openpyxl"
        )

    wb = Workbook()
    # Remove the default sheet - we'll create our own.
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    label_font = Font(bold=True)
    money_fmt = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
    pct_fmt = '0.00%'
    plain_pct_fmt = '0.00"%"'  # for inputs that the user types as a percent number

    # =========================================================
    # Sheet 1 — Assumptions (the only sheet with raw input numbers)
    # =========================================================
    ws_a = wb.create_sheet("Assumptions")
    ws_a["A1"] = "Assumption"
    ws_a["B1"] = "Value"
    for col in ("A1", "B1"):
        ws_a[col].font = header_font
        ws_a[col].fill = header_fill

    # Row layout — keep this map; the formulas below reference these row numbers.
    assumption_rows = [
        ("Starting revenue", inputs.revenue_0, money_fmt),                           # row 2
        ("Year 1 growth (%)", inputs.growth_years_1_5[0] * 100, plain_pct_fmt),      # row 3
        ("Year 2 growth (%)", inputs.growth_years_1_5[1] * 100, plain_pct_fmt),      # row 4
        ("Year 3 growth (%)", inputs.growth_years_1_5[2] * 100, plain_pct_fmt),      # row 5
        ("Year 4 growth (%)", inputs.growth_years_1_5[3] * 100, plain_pct_fmt),      # row 6
        ("Year 5 growth (%)", inputs.growth_years_1_5[4] * 100, plain_pct_fmt),      # row 7
        ("Current EBIT margin (%)", inputs.ebit_margin_0 * 100, plain_pct_fmt),      # row 8
        ("Target EBIT margin (%)", inputs.ebit_margin_target * 100, plain_pct_fmt),  # row 9
        ("Tax rate (%)", inputs.tax_rate * 100, plain_pct_fmt),                      # row 10
        ("D&A as % of revenue", inputs.da_pct_revenue * 100, plain_pct_fmt),         # row 11
        ("Capex as % of revenue", inputs.capex_pct_revenue * 100, plain_pct_fmt),    # row 12
        ("NWC as % of revenue", inputs.nwc_pct_revenue * 100, plain_pct_fmt),        # row 13
        ("WACC (%)", inputs.wacc * 100, plain_pct_fmt),                              # row 14
        ("Terminal growth (%)", inputs.terminal_growth * 100, plain_pct_fmt),        # row 15
        ("Shares outstanding", inputs.shares_outstanding, '#,##0'),                  # row 16
        ("Net debt", inputs.net_debt, money_fmt),                                    # row 17
    ]
    for i, (label, value, fmt) in enumerate(assumption_rows, start=2):
        ws_a.cell(row=i, column=1, value=label).font = label_font
        cell = ws_a.cell(row=i, column=2, value=value)
        cell.number_format = fmt

    ws_a.column_dimensions["A"].width = 32
    ws_a.column_dimensions["B"].width = 22

    # Named cell references for readability in formulas below.
    A_REV   = "Assumptions!$B$2"
    A_G    = ["Assumptions!$B$3", "Assumptions!$B$4", "Assumptions!$B$5",
              "Assumptions!$B$6", "Assumptions!$B$7"]
    A_EM0   = "Assumptions!$B$8"
    A_EMT   = "Assumptions!$B$9"
    A_TAX   = "Assumptions!$B$10"
    A_DA    = "Assumptions!$B$11"
    A_CAPEX = "Assumptions!$B$12"
    A_NWC   = "Assumptions!$B$13"
    A_WACC  = "Assumptions!$B$14"
    A_TG    = "Assumptions!$B$15"
    A_SH    = "Assumptions!$B$16"
    A_ND    = "Assumptions!$B$17"

    # =========================================================
    # Sheet 2 — Forecast (rows 2..6 are years 1..5, all formulas)
    # =========================================================
    ws_f = wb.create_sheet("Forecast")
    forecast_headers = [
        "Year", "Growth %", "Revenue", "EBIT Margin %", "EBIT",
        "NOPAT", "D&A", "Capex", "NWC", "Change in NWC",
        "FCFF", "Discount Factor", "PV of FCFF",
    ]
    for col_idx, h in enumerate(forecast_headers, start=1):
        cell = ws_f.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Years 1..5 → rows 2..6
    for year in range(1, 6):
        r = year + 1  # row in sheet
        prev_r = r - 1  # previous year's row (for r=2 this is the header row, handled below)

        # Column A — Year number
        ws_f.cell(row=r, column=1, value=year)

        # Column B — Growth rate (pulls Y1..Y5 from Assumptions)
        ws_f.cell(row=r, column=2, value=f"={A_G[year - 1]}").number_format = plain_pct_fmt

        # Column C — Revenue
        if year == 1:
            # First-year revenue grows off of the starting revenue input.
            ws_f.cell(row=r, column=3, value=f"={A_REV}*(1+B{r}/100)")
        else:
            ws_f.cell(row=r, column=3, value=f"=C{prev_r}*(1+B{r}/100)")
        ws_f.cell(row=r, column=3).number_format = money_fmt

        # Column D — EBIT margin (linearly interpolates from current to target by Y5)
        ws_f.cell(row=r, column=4,
                  value=f"={A_EM0}+({A_EMT}-{A_EM0})*(A{r}/5)").number_format = plain_pct_fmt

        # Column E — EBIT = Revenue × EBIT margin
        ws_f.cell(row=r, column=5, value=f"=C{r}*D{r}/100").number_format = money_fmt

        # Column F — NOPAT = EBIT × (1 − tax rate)
        ws_f.cell(row=r, column=6, value=f"=E{r}*(1-{A_TAX}/100)").number_format = money_fmt

        # Column G — D&A
        ws_f.cell(row=r, column=7, value=f"=C{r}*{A_DA}/100").number_format = money_fmt

        # Column H — Capex
        ws_f.cell(row=r, column=8, value=f"=C{r}*{A_CAPEX}/100").number_format = money_fmt

        # Column I — NWC
        ws_f.cell(row=r, column=9, value=f"=C{r}*{A_NWC}/100").number_format = money_fmt

        # Column J — Change in NWC (Y1 uses Starting Revenue × NWC%, later years use prior I)
        if year == 1:
            ws_f.cell(row=r, column=10, value=f"=I{r}-{A_REV}*{A_NWC}/100").number_format = money_fmt
        else:
            ws_f.cell(row=r, column=10, value=f"=I{r}-I{prev_r}").number_format = money_fmt

        # Column K — FCFF = NOPAT + D&A − Capex − ΔNWC
        ws_f.cell(row=r, column=11, value=f"=F{r}+G{r}-H{r}-J{r}").number_format = money_fmt

        # Column L — Discount factor = 1 / (1+WACC)^t
        ws_f.cell(row=r, column=12, value=f"=1/(1+{A_WACC}/100)^A{r}").number_format = '0.0000'

        # Column M — PV of FCFF
        ws_f.cell(row=r, column=13, value=f"=K{r}*L{r}").number_format = money_fmt

    # Set sensible column widths
    widths = [6, 10, 18, 14, 18, 18, 14, 14, 16, 16, 18, 14, 18]
    for i, w in enumerate(widths, start=1):
        ws_f.column_dimensions[get_column_letter(i)].width = w

    # =========================================================
    # Sheet 3 — DCF Bridge (entirely formulas)
    # =========================================================
    ws_b = wb.create_sheet("DCF Bridge")
    ws_b["A1"] = "Component"
    ws_b["B1"] = "Value"
    for col in ("A1", "B1"):
        ws_b[col].font = header_font
        ws_b[col].fill = header_fill

    bridge_rows = [
        # row, label, formula, format
        (2, "PV of forecast FCFF", "=SUM(Forecast!M2:M6)", money_fmt),
        # Terminal value formula: FCFF(Y5)*(1+g)/(WACC-g) / (1+WACC)^5
        (3, "PV of terminal value",
            f"=Forecast!K6*(1+{A_TG}/100)/({A_WACC}/100-{A_TG}/100)/(1+{A_WACC}/100)^5",
            money_fmt),
        (4, "Enterprise value", "=B2+B3", money_fmt),
        (5, "Less: Net debt", f"=-{A_ND}", money_fmt),
        (6, "Equity value", "=B4+B5", money_fmt),
        (7, "Shares outstanding", f"={A_SH}", '#,##0'),
        (8, "Intrinsic value / share", "=B6/B7", money_fmt),
    ]
    for r, label, formula, fmt in bridge_rows:
        ws_b.cell(row=r, column=1, value=label).font = label_font
        cell = ws_b.cell(row=r, column=2, value=formula)
        cell.number_format = fmt
    # Highlight the final per-share row
    ws_b.cell(row=8, column=1).font = Font(bold=True)
    ws_b.cell(row=8, column=2).font = Font(bold=True)
    ws_b.column_dimensions["A"].width = 26
    ws_b.column_dimensions["B"].width = 22

    # =========================================================
    # Sheet 4 — Summary (mostly references the other sheets)
    # =========================================================
    ws_s = wb.create_sheet("Summary")
    ws_s["A1"] = "Metric"
    ws_s["B1"] = "Value"
    for col in ("A1", "B1"):
        ws_s[col].font = header_font
        ws_s[col].fill = header_fill

    # Current price is a market observation, not part of the DCF chain,
    # so it stays as a literal value.
    market_price = inputs.current_price if inputs.current_price > 0 else None

    summary_rows = [
        (2, "Ticker", inputs.ticker, "@"),
        (3, "Company", inputs.company_name, "@"),
        (4, "Current market price", market_price, money_fmt),
        (5, "Intrinsic value / share", "='DCF Bridge'!B8", money_fmt),
        # Upside formula refers to row 5 above and row 4 (market price)
        (6, "Upside / downside (%)", "=IF(B4>0,(B5/B4-1)*100,\"N/A\")", '0.00"%"'),
        (7, "Enterprise value", "='DCF Bridge'!B4", money_fmt),
        (8, "Equity value", "='DCF Bridge'!B6", money_fmt),
        (9, "Shares outstanding", f"={A_SH}", '#,##0'),
        (10, "Net debt", f"={A_ND}", money_fmt),
    ]
    for r, label, value, fmt in summary_rows:
        ws_s.cell(row=r, column=1, value=label).font = label_font
        cell = ws_s.cell(row=r, column=2, value=value)
        cell.number_format = fmt
    ws_s.column_dimensions["A"].width = 28
    ws_s.column_dimensions["B"].width = 22

    # =========================================================
    # Sheet 5 — Sensitivity (static values; a what-if grid)
    # =========================================================
    ws_v = wb.create_sheet("Sensitivity")
    ws_v["A1"] = "Terminal Growth \\ WACC"
    ws_v["A1"].font = header_font
    ws_v["A1"].fill = header_fill

    # Column headers (WACC values)
    for j, w in enumerate(sensitivity_df.columns, start=2):
        cell = ws_v.cell(row=1, column=j, value=f"{w * 100:.1f}%")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Row headers (terminal growth values) and data
    for i, tg in enumerate(sensitivity_df.index, start=2):
        ws_v.cell(row=i, column=1, value=f"{tg * 100:.1f}%").font = label_font
        for j, w in enumerate(sensitivity_df.columns, start=2):
            val = sensitivity_df.loc[tg, w]
            if pd.isna(val):
                ws_v.cell(row=i, column=j, value="—")
            else:
                cell = ws_v.cell(row=i, column=j, value=float(val))
                cell.number_format = money_fmt

    ws_v.column_dimensions["A"].width = 24
    for j in range(2, len(sensitivity_df.columns) + 2):
        ws_v.column_dimensions[get_column_letter(j)].width = 14

    # =========================================================
    # Sheet 6 — Read Me
    # =========================================================
    ws_r = wb.create_sheet("Read Me")
    ws_r["A1"] = "How this workbook works"
    ws_r["A1"].font = Font(bold=True, size=14)
    readme_lines = [
        "",
        "• The Assumptions sheet is the only place with raw input numbers.",
        "• The Forecast, DCF Bridge, and Summary sheets are built from formulas",
        "  that reference Assumptions. Edit any assumption and every downstream",
        "  number recalculates automatically.",
        "",
        "• Try it: change WACC on the Assumptions sheet (cell B14) and watch",
        "  the intrinsic value per share on the Summary sheet update.",
        "",
        "• The Sensitivity sheet is a static what-if grid generated when the",
        "  workbook was exported. To refresh it, re-export from the app.",
        "",
        "Formula reference (all driven from Assumptions):",
        "• Revenue(t)   = Revenue(t-1) × (1 + Growth Rate)",
        "• EBIT Margin  = current + (target − current) × (year / 5)",
        "• EBIT         = Revenue × EBIT Margin",
        "• NOPAT        = EBIT × (1 − Tax Rate)",
        "• FCFF         = NOPAT + D&A − Capex − Change in NWC",
        "• PV of FCFF   = FCFF / (1 + WACC)^t",
        "• Terminal V.  = FCFF(Y5) × (1 + g) / (WACC − g)",
        "• PV of TV     = Terminal Value / (1 + WACC)^5",
        "• EV           = SUM(PV of FCFF) + PV of TV",
        "• Equity Value = EV − Net Debt",
        "• Per share    = Equity Value / Shares Outstanding",
    ]
    for i, line in enumerate(readme_lines, start=2):
        ws_r.cell(row=i, column=1, value=line)
    ws_r.column_dimensions["A"].width = 80

    # Reorder so Summary is the first sheet the user sees
    wb.move_sheet("Summary", offset=-(len(wb.sheetnames) - 1))

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# =========================================================
# Navigation state
# =========================================================
PAGES = ["Home", "Company Data", "Assumptions", "Valuation",
         "Step-by-Step Breakdown", "Sensitivity", "Download"]

st.session_state.setdefault("page", "Home")
st.session_state.setdefault("stock_data", None)
st.session_state.setdefault("last_ticker", "")
st.session_state.setdefault("assumptions", {
    "g1": 10.0, "g2": 8.0, "g3": 6.0, "g4": 5.0, "g5": 4.0,
    "ebit_margin_target": 18.0,
    "tax_rate": 25.0,
    "da_pct": 3.0,
    "capex_pct": 4.0,
    "wacc": 9.0,
    "terminal_growth": 2.5,
})


def go_to(page: str):
    st.session_state["page"] = page


def render_nav_bar():
    """Compact pill nav at the top of every non-Home page."""
    cols = st.columns(len(PAGES))
    for i, p in enumerate(PAGES):
        is_current = st.session_state["page"] == p
        btn_type = "primary" if is_current else "secondary"
        if cols[i].button(p, key=f"nav_{p}", type=btn_type, use_container_width=True):
            go_to(p)
            st.rerun()
    st.divider()


# =========================================================
# Sidebar: ticker selection (always visible)
# =========================================================
with st.sidebar:
    st.header("Company")
    ticker = st.text_input("Ticker symbol", value="AAPL").upper().strip()
    fetch_data = st.button("Load company data", use_container_width=True)

    if fetch_data:
        st.session_state["stock_data"] = get_stock_data(ticker)
        st.session_state["last_ticker"] = ticker

    if st.session_state["stock_data"] is None:
        st.session_state["stock_data"] = get_stock_data(ticker)
        st.session_state["last_ticker"] = ticker

    sd = st.session_state.get("stock_data")
    if sd and sd.get("success"):
        st.success(f"Loaded: {sd.get('company_name', '')}")
    elif sd:
        st.error(sd.get("error", "Could not load data."))

    st.divider()
    st.caption("Use the navigation buttons on the home page or at the top of each page to move around.")


# =========================================================
# Verify data loaded before continuing
# =========================================================
stock_data = st.session_state.get("stock_data")
if stock_data is None or not isinstance(stock_data, dict) or not stock_data.get("success"):
    st.title("Equity Valuation App")
    st.error("Could not load company data. Use the sidebar to enter a ticker and click 'Load company data'.")
    if isinstance(stock_data, dict) and stock_data.get("diagnostics"):
        with st.expander("Diagnostics"):
            for d in stock_data["diagnostics"]:
                st.code(d, language=None)
    st.stop()

loaded_ticker = stock_data.get("ticker", "")
company_name = stock_data.get("company_name", loaded_ticker or ticker)
current_price = stock_data.get("current_price", np.nan)
revenue_auto = stock_data.get("revenue", np.nan)
shares_auto = stock_data.get("shares_outstanding", np.nan)
net_debt_auto = stock_data.get("net_debt", np.nan)
ebit_margin_auto = stock_data.get("ebit_margin", np.nan)
nwc_pct_auto = stock_data.get("nwc_pct_revenue", np.nan)


def build_inputs_from_state() -> DCFInputs:
    """Combine auto-fetched company data with user-controlled assumptions."""
    a = st.session_state["assumptions"]

    rev = revenue_auto if not pd.isna(revenue_auto) else 0.0
    shares = shares_auto if not pd.isna(shares_auto) and shares_auto > 0 else 1.0
    nd = net_debt_auto if not pd.isna(net_debt_auto) else 0.0
    price = current_price if not pd.isna(current_price) else 0.0

    # Current EBIT margin: use auto value if we have one, otherwise start at the target.
    if not pd.isna(ebit_margin_auto):
        em0 = ebit_margin_auto
    else:
        em0 = pct_from_ui(a["ebit_margin_target"])

    # NWC%: use auto if available, else 8% generic default.
    nwc_pct = nwc_pct_auto if not pd.isna(nwc_pct_auto) else 0.08

    return DCFInputs(
        ticker=loaded_ticker,
        company_name=company_name,
        revenue_0=rev,
        growth_years_1_5=[
            pct_from_ui(a["g1"]), pct_from_ui(a["g2"]), pct_from_ui(a["g3"]),
            pct_from_ui(a["g4"]), pct_from_ui(a["g5"]),
        ],
        ebit_margin_0=em0,
        ebit_margin_target=pct_from_ui(a["ebit_margin_target"]),
        tax_rate=pct_from_ui(a["tax_rate"]),
        da_pct_revenue=pct_from_ui(a["da_pct"]),
        capex_pct_revenue=pct_from_ui(a["capex_pct"]),
        nwc_pct_revenue=nwc_pct,
        wacc=pct_from_ui(a["wacc"]),
        terminal_growth=pct_from_ui(a["terminal_growth"]),
        shares_outstanding=shares,
        net_debt=nd,
        current_price=price,
    )


# =========================================================
# PAGE: Home
# =========================================================
def page_home():
    st.title("Equity Valuation App")
    st.caption(f"DCF model for {company_name} ({loaded_ticker}) — pulls live market data and lets you control valuation assumptions.")

    st.markdown("### Where would you like to go?")
    st.write("")

    button_specs = [
        ("📊 Company Data", "Company Data",
         "See the financials and market data pulled from Yahoo Finance."),
        ("⚙️ Assumptions", "Assumptions",
         "Choose growth rates, margins, WACC, and other forecast inputs."),
        ("💰 Valuation", "Valuation",
         "View intrinsic value per share and the comparison to market price."),
        ("📋 Step-by-Step Breakdown", "Step-by-Step Breakdown",
         "See each step of the DCF calculation, from revenue to per-share value."),
        ("📈 Sensitivity", "Sensitivity",
         "Test how the per-share value changes with WACC and terminal growth."),
        ("📥 Download", "Download",
         "Export the full valuation to a multi-sheet Excel workbook."),
    ]

    for i in range(0, len(button_specs), 2):
        cols = st.columns(2)
        for j, col in enumerate(cols):
            if i + j < len(button_specs):
                label, target, desc = button_specs[i + j]
                with col:
                    with st.container(border=True):
                        st.markdown(f"#### {label}")
                        st.caption(desc)
                        if st.button(f"Go to {target}", key=f"home_{target}", use_container_width=True):
                            go_to(target)
                            st.rerun()

    st.divider()
    with st.expander("What this app does"):
        st.markdown(
            """
            - Enter a **stock ticker**
            - Automatically pull market and financial data from Yahoo Finance
            - Choose the **forecast assumptions** that drive the DCF model
            - Compute **intrinsic value** per share using a five-year DCF plus terminal value
            - Compare intrinsic value to the **current stock price**
            - See the **intermediate steps** so the valuation is transparent and reproducible in Excel
            """
        )


# =========================================================
# PAGE: Company Data
# =========================================================
def page_company_data():
    render_nav_bar()
    st.title("Company Data")
    st.caption("These values come directly from Yahoo Finance and feed the DCF model. They are not editable — change the ticker in the sidebar to load a different company.")

    sector = stock_data.get("sector", "N/A")
    industry = stock_data.get("industry", "N/A")
    if sector != "N/A" or industry != "N/A":
        st.markdown(f"**Sector:** {sector}  ·  **Industry:** {industry}")

    cols = st.columns(4)
    cols[0].metric("Current price", fmt_currency(current_price))
    cols[1].metric("Shares outstanding", fmt_number(shares_auto))
    cols[2].metric("Latest revenue", fmt_currency(revenue_auto))
    cols[3].metric("Net debt", fmt_currency(net_debt_auto))

    cols2 = st.columns(4)
    cols2[0].metric("Current EBIT margin", fmt_pct(ebit_margin_auto))
    cols2[1].metric("Working capital ratio", fmt_pct(nwc_pct_auto))
    cols2[2].metric("Market cap", fmt_currency(stock_data.get("market_cap", np.nan)))
    cols2[3].metric("Total debt", fmt_currency(stock_data.get("total_debt", np.nan)))

    st.divider()
    st.subheader("1-year price history")
    history = stock_data.get("price_history")
    if (
        isinstance(history, pd.DataFrame)
        and not history.empty
        and "Close" in history.columns
    ):
        price_df = history.reset_index()
        date_col = next((c for c in ("Date", "Datetime", "index") if c in price_df.columns), None)
        if date_col is not None:
            fig_price = px.line(price_df, x=date_col, y="Close", title=f"{loaded_ticker} closing price")
            st.plotly_chart(fig_price, use_container_width=True)
        else:
            st.info("Price history is available but the date column could not be identified.")
    else:
        st.info("Price history is unavailable for this ticker.")

    missing = []
    if pd.isna(revenue_auto): missing.append("Revenue")
    if pd.isna(ebit_margin_auto): missing.append("EBIT margin")
    if pd.isna(nwc_pct_auto): missing.append("Working capital ratio")
    if pd.isna(net_debt_auto): missing.append("Net debt")
    if pd.isna(shares_auto): missing.append("Shares outstanding")
    if missing:
        st.warning(
            f"The following fields could not be retrieved: **{', '.join(missing)}**. "
            "The model will use neutral defaults for any missing values."
        )
        diags = stock_data.get("diagnostics", [])
        if diags:
            with st.expander("Diagnostic details"):
                for d in diags:
                    st.code(d, language=None)


# =========================================================
# PAGE: Assumptions
# =========================================================
def page_assumptions():
    render_nav_bar()
    st.title("Assumptions")
    st.caption(
        "These are the valuation choices you make. Each section shows historical "
        "benchmarks where available — your assumption doesn't have to match, but "
        "deviations should be deliberate."
    )

    a = st.session_state["assumptions"]

    # Pull historical benchmarks from the cached company data
    hist_growth = stock_data.get("hist_revenue_growth_avg", np.nan)
    hist_growth_list = stock_data.get("hist_revenue_growth_list", [])
    hist_em_mean = stock_data.get("hist_ebit_margin_mean", np.nan)
    hist_em_min = stock_data.get("hist_ebit_margin_min", np.nan)
    hist_em_max = stock_data.get("hist_ebit_margin_max", np.nan)
    hist_da = stock_data.get("hist_da_pct", np.nan)
    hist_capex = stock_data.get("hist_capex_pct", np.nan)
    hist_nwc = stock_data.get("nwc_pct_revenue", np.nan)

    # Macro benchmarks (constants — reasonable approximations for US/global)
    NOMINAL_GDP_GROWTH = 4.5      # Long-run US nominal GDP growth, used as growth ceiling
    REAL_GDP_GROWTH = 2.0         # Long-run real GDP growth
    TYPICAL_TAX_LOW, TYPICAL_TAX_HIGH = 18.0, 28.0
    TYPICAL_WACC_LOW, TYPICAL_WACC_HIGH = 6.0, 12.0

    # =========================================================
    # Revenue growth section
    # =========================================================
    st.subheader("Revenue growth (% per year)")
    st.caption("Year-by-year revenue growth used to project sales over the next five years.")

    # Benchmark info box
    with st.container(border=True):
        cols = st.columns(2)
        with cols[0]:
            st.markdown("**Historical revenue growth**")
            if not pd.isna(hist_growth):
                st.markdown(f"3-year average: **{hist_growth * 100:.2f}%**")
                if hist_growth_list:
                    yoy_str = ", ".join(f"{x * 100:.1f}%" for x in hist_growth_list)
                    st.caption(f"Year-over-year: {yoy_str}")
            else:
                st.caption("Historical growth is unavailable for this company.")
        with cols[1]:
            st.markdown("**Long-run ceiling**")
            st.markdown(f"Nominal GDP growth ≈ **{NOMINAL_GDP_GROWTH:.1f}%**")
            st.caption(
                "By Year 5, growth should be approaching this ceiling. "
                "No company can grow faster than the economy forever."
            )

    g_cols = st.columns(5)
    a["g1"] = g_cols[0].number_input("Year 1", value=a["g1"], step=0.5, key="g1_in")
    a["g2"] = g_cols[1].number_input("Year 2", value=a["g2"], step=0.5, key="g2_in")
    a["g3"] = g_cols[2].number_input("Year 3", value=a["g3"], step=0.5, key="g3_in")
    a["g4"] = g_cols[3].number_input("Year 4", value=a["g4"], step=0.5, key="g4_in")
    a["g5"] = g_cols[4].number_input("Year 5", value=a["g5"], step=0.5, key="g5_in")

    # Validation: Year 5 should not exceed nominal GDP growth
    if a["g5"] > NOMINAL_GDP_GROWTH:
        st.warning(
            f"⚠️ Year 5 growth of **{a['g5']:.1f}%** exceeds long-run nominal GDP growth "
            f"(~{NOMINAL_GDP_GROWTH:.1f}%). Few companies can sustain above-GDP growth "
            "for five years and beyond — consider tapering Year 5 lower."
        )
    # Sanity: also flag if any year is unusually high (>30%) or negative for >2 years
    if max(a["g1"], a["g2"], a["g3"], a["g4"], a["g5"]) > 30.0:
        st.info("ℹ️ Some growth rates exceed 30%. That's plausible for very young or hyper-growth companies, but unusual for mature businesses.")

    st.divider()

    # =========================================================
    # Operating assumptions section
    # =========================================================
    st.subheader("Operating assumptions")

    # EBIT margin benchmark
    with st.container(border=True):
        st.markdown("**Historical EBIT margin**")
        if not pd.isna(hist_em_mean):
            cols = st.columns(3)
            cols[0].metric("Average", f"{hist_em_mean * 100:.2f}%")
            cols[1].metric("Min", f"{hist_em_min * 100:.2f}%")
            cols[2].metric("Max", f"{hist_em_max * 100:.2f}%")
            st.caption(
                "Your target EBIT margin should be grounded in this range. "
                "A target much higher than the historical max implies meaningful margin expansion — "
                "consider whether the company has a credible path there."
            )
        else:
            st.caption("Historical EBIT margins are unavailable for this company.")

    op_cols = st.columns(2)
    with op_cols[0]:
        a["ebit_margin_target"] = st.number_input(
            "Target EBIT margin (%)",
            value=a["ebit_margin_target"], step=0.5, key="emt_in",
            help="The long-run EBIT margin the company reaches by Year 5. The starting margin comes from company data.",
        )
        # Warning if target is way above historical max
        if not pd.isna(hist_em_max) and a["ebit_margin_target"] > hist_em_max * 100 + 5:
            st.warning(
                f"⚠️ Target margin **{a['ebit_margin_target']:.1f}%** is more than 5 points above "
                f"the historical max of **{hist_em_max * 100:.1f}%**."
            )
        elif not pd.isna(hist_em_min) and a["ebit_margin_target"] < hist_em_min * 100 - 5:
            st.warning(
                f"⚠️ Target margin **{a['ebit_margin_target']:.1f}%** is more than 5 points below "
                f"the historical min of **{hist_em_min * 100:.1f}%**."
            )

        a["tax_rate"] = st.number_input(
            "Tax rate (%)",
            value=a["tax_rate"], step=0.5, key="tax_in",
            help=f"Corporate tax rate applied to EBIT to estimate NOPAT. Typical effective rate: {TYPICAL_TAX_LOW:.0f}–{TYPICAL_TAX_HIGH:.0f}%.",
        )
        st.caption(f"💡 US statutory rate: 21%. Typical effective rate after deductions: {TYPICAL_TAX_LOW:.0f}–{TYPICAL_TAX_HIGH:.0f}%.")
        if a["tax_rate"] < TYPICAL_TAX_LOW or a["tax_rate"] > TYPICAL_TAX_HIGH + 10:
            st.warning(f"⚠️ Tax rate **{a['tax_rate']:.1f}%** is outside the typical {TYPICAL_TAX_LOW:.0f}–{TYPICAL_TAX_HIGH:.0f}% range.")

    with op_cols[1]:
        a["da_pct"] = st.number_input(
            "D&A as % of revenue",
            value=a["da_pct"], step=0.2, key="da_in",
            help="Depreciation & amortization, added back in the FCFF calculation.",
        )
        if not pd.isna(hist_da):
            st.caption(f"💡 This company's historical D&A: **{hist_da * 100:.2f}%** of revenue (multi-year average).")
        else:
            st.caption("Typical range varies widely by industry (often 2–8% for asset-light businesses, higher for capital-intensive ones).")

        a["capex_pct"] = st.number_input(
            "Capex as % of revenue",
            value=a["capex_pct"], step=0.2, key="capex_in",
            help="Capital expenditures required to support future growth.",
        )
        if not pd.isna(hist_capex):
            st.caption(f"💡 This company's historical capex: **{hist_capex * 100:.2f}%** of revenue (multi-year average).")
        else:
            st.caption("Typical range varies widely by industry (often 3–6% for software/services, 8–15% for manufacturing/utilities).")

    st.divider()

    # =========================================================
    # Discount rate / terminal value section
    # =========================================================
    st.subheader("Discount rate and terminal value")

    with st.container(border=True):
        cols = st.columns(2)
        with cols[0]:
            st.markdown("**Typical WACC range**")
            st.markdown(f"Most large-cap companies: **{TYPICAL_WACC_LOW:.0f}–{TYPICAL_WACC_HIGH:.0f}%**")
            st.caption(
                "Lower WACC = lower perceived risk (e.g. utilities, blue-chips). "
                "Higher WACC = higher risk (e.g. early-stage tech, cyclicals)."
            )
        with cols[1]:
            st.markdown("**Terminal growth ceiling**")
            st.markdown(f"Long-run nominal GDP: **~{NOMINAL_GDP_GROWTH:.1f}%**")
            st.markdown(f"Long-run real GDP: **~{REAL_GDP_GROWTH:.1f}%**")
            st.caption(
                "Terminal growth represents perpetual growth after Year 5. "
                "It cannot exceed the long-run growth of the economy."
            )

    dr_cols = st.columns(2)
    with dr_cols[0]:
        a["wacc"] = st.number_input(
            "WACC (%)",
            min_value=0.1, value=a["wacc"], step=0.25, key="wacc_in",
            help="Weighted average cost of capital used to discount FCFF.",
        )
        if a["wacc"] < TYPICAL_WACC_LOW or a["wacc"] > TYPICAL_WACC_HIGH:
            st.warning(
                f"⚠️ WACC **{a['wacc']:.2f}%** is outside the typical "
                f"{TYPICAL_WACC_LOW:.0f}–{TYPICAL_WACC_HIGH:.0f}% range for large-cap companies."
            )

    with dr_cols[1]:
        a["terminal_growth"] = st.number_input(
            "Terminal growth (%)",
            value=a["terminal_growth"], step=0.25, key="tg_in",
            help="Perpetual growth rate after Year 5. Must be below WACC and below long-run GDP growth.",
        )
        if a["terminal_growth"] > NOMINAL_GDP_GROWTH:
            st.error(
                f"⚠️ Terminal growth **{a['terminal_growth']:.2f}%** exceeds long-run nominal GDP "
                f"growth (~{NOMINAL_GDP_GROWTH:.1f}%). No company can grow faster than the economy "
                "forever — this assumption is mathematically allowable but economically unrealistic."
            )
        elif a["terminal_growth"] < 0:
            st.info(
                f"ℹ️ Negative terminal growth implies the business shrinks forever. "
                "Reasonable for declining industries, unusual otherwise."
            )

    if pct_from_ui(a["terminal_growth"]) >= pct_from_ui(a["wacc"]):
        st.error("Terminal growth must be lower than WACC for the Gordon Growth formula to work.")

    st.divider()
    st.info(
        "When you're done setting assumptions, click **Valuation** in the navigation "
        "above to see the result, or **Step-by-Step Breakdown** to see how each "
        "number flows through the model."
    )


# =========================================================
# PAGE: Valuation
# =========================================================
def page_valuation():
    render_nav_bar()
    st.title("Valuation")

    inputs = build_inputs_from_state()
    if inputs.terminal_growth >= inputs.wacc:
        st.error("Terminal growth must be lower than WACC. Edit assumptions to fix this.")
        return
    if inputs.revenue_0 <= 0:
        st.error("Starting revenue (from company data) is zero or unavailable, so the DCF cannot run.")
        return

    results = run_dcf(inputs)

    metric_cols = st.columns(5)
    metric_cols[0].metric("Enterprise value", fmt_currency(results["enterprise_value"]))
    metric_cols[1].metric("Equity value", fmt_currency(results["equity_value"]))
    metric_cols[2].metric("Intrinsic value / share", fmt_currency(results["value_per_share"]))
    metric_cols[3].metric(
        "Current market price",
        fmt_currency(inputs.current_price) if inputs.current_price > 0 else "N/A",
    )
    metric_cols[4].metric(
        "Upside / downside",
        "N/A" if pd.isna(results["premium_discount"]) else f"{results['premium_discount'] * 100:,.2f}%",
    )

    st.divider()
    if (
        not pd.isna(results["value_per_share"])
        and inputs.current_price > 0
        and results["equity_value"] > 0
    ):
        if results["value_per_share"] > inputs.current_price:
            st.success("Based on your assumptions, the stock appears **undervalued** — intrinsic value is above the market price.")
        elif results["value_per_share"] < inputs.current_price:
            st.warning("Based on your assumptions, the stock appears **overvalued** — intrinsic value is below the market price.")
        else:
            st.info("Based on your assumptions, intrinsic value is approximately equal to the market price.")
    elif results["equity_value"] <= 0:
        st.warning(
            "Equity value is zero or negative — net debt exceeds the discounted value "
            "of future cash flows. Try more optimistic growth, margin, or WACC assumptions."
        )
    elif inputs.current_price <= 0:
        st.info("Current market price is unavailable, so no comparison can be shown.")

    st.divider()
    st.subheader("Operating forecast")
    margin_df = results["forecast"][["Year", "Revenue", "FCFF"]].melt(
        id_vars="Year", var_name="Metric", value_name="Value"
    )
    fig_ops = px.line(margin_df, x="Year", y="Value", color="Metric", markers=True)
    st.plotly_chart(fig_ops, use_container_width=True)


# =========================================================
# PAGE: Step-by-Step Breakdown
# =========================================================
def page_breakdown():
    render_nav_bar()
    st.title("Step-by-Step DCF Breakdown")
    st.caption("Each step of the calculation, transparent and reproducible.")

    inputs = build_inputs_from_state()
    if inputs.terminal_growth >= inputs.wacc:
        st.error("Terminal growth must be lower than WACC. Edit assumptions to fix this.")
        return
    if inputs.revenue_0 <= 0:
        st.error("Starting revenue is zero or unavailable, so the DCF cannot run.")
        return

    results = run_dcf(inputs)
    forecast = results["forecast"]
    discounted = results["discounted"]

    with st.expander("Step 1 — Forecast revenue", expanded=True):
        st.markdown("**Revenue(t) = Revenue(t-1) × (1 + Growth Rate)**")
        rev_view = forecast[["Year", "Revenue", "Revenue Growth"]].copy()
        rev_view["Revenue Growth"] = (rev_view["Revenue Growth"] * 100).map(lambda x: f"{x:.2f}%")
        st.dataframe(rev_view, use_container_width=True)

    with st.expander("Step 2 — Forecast operating profit"):
        st.markdown("**EBIT = Revenue × EBIT Margin** (margin moves from current to target by Year 5)")
        op_view = forecast[["Year", "Revenue", "EBIT Margin", "EBIT"]].copy()
        op_view["EBIT Margin"] = (op_view["EBIT Margin"] * 100).map(lambda x: f"{x:.2f}%")
        st.dataframe(op_view, use_container_width=True)

    with st.expander("Step 3 — After-tax operating income (NOPAT)"):
        st.markdown("**NOPAT = EBIT × (1 − Tax Rate)**")
        st.dataframe(forecast[["Year", "EBIT", "NOPAT"]], use_container_width=True)

    with st.expander("Step 4 — Free cash flow to the firm (FCFF)"):
        st.markdown("**FCFF = NOPAT + D&A − Capex − Change in NWC**")
        st.dataframe(forecast[["Year", "NOPAT", "D&A", "Capex", "Change in NWC", "FCFF"]], use_container_width=True)

    with st.expander("Step 5 — Discount FCFF to present value"):
        st.markdown("**PV of FCFF = FCFF / (1 + WACC)^t**")
        st.dataframe(discounted[["Year", "FCFF", "Discount Factor", "PV of FCFF"]], use_container_width=True)

    with st.expander("Step 6 — Estimate terminal value"):
        st.markdown("**Terminal Value = FCFF(Year 5) × (1 + g) / (WACC − g)**")
        tv_df = pd.DataFrame({
            "Metric": ["FCFF in Year 5", "Terminal growth", "WACC", "Terminal value", "PV of terminal value"],
            "Value": [
                forecast.iloc[-1]["FCFF"],
                f"{inputs.terminal_growth * 100:.2f}%",
                f"{inputs.wacc * 100:.2f}%",
                results["terminal_value"],
                results["pv_terminal_value"],
            ],
        })
        st.dataframe(tv_df, use_container_width=True)

    with st.expander("Step 7 — Enterprise value, equity value, and value per share"):
        st.markdown("**Equity Value = Enterprise Value − Net Debt**  \n**Intrinsic Value / Share = Equity Value / Shares Outstanding**")
        bridge_df = pd.DataFrame({
            "Component": [
                "PV of forecast FCFF", "PV of terminal value", "Enterprise value",
                "Less: Net debt", "Equity value", "Shares outstanding", "Intrinsic value / share",
            ],
            "Value": [
                results["pv_fcf"], results["pv_terminal_value"], results["enterprise_value"],
                -inputs.net_debt, results["equity_value"], inputs.shares_outstanding,
                results["value_per_share"],
            ],
        })
        st.dataframe(bridge_df, use_container_width=True)

    st.divider()
    st.subheader("Full forecast table")
    forecast_display = discounted.copy()
    for col in ["Revenue Growth", "EBIT Margin"]:
        forecast_display[col] = (forecast_display[col] * 100).map(lambda x: f"{x:.2f}%")
    st.dataframe(forecast_display, use_container_width=True)


# =========================================================
# PAGE: Sensitivity
# =========================================================
def page_sensitivity():
    render_nav_bar()
    st.title("Sensitivity Analysis")
    st.caption("How the per-share value changes when WACC and terminal growth move.")

    inputs = build_inputs_from_state()
    if inputs.terminal_growth >= inputs.wacc:
        st.error("Terminal growth must be lower than WACC. Edit assumptions to fix this.")
        return
    if inputs.revenue_0 <= 0:
        st.error("Starting revenue is zero or unavailable, so the DCF cannot run.")
        return

    sensitivity = make_sensitivity_table(inputs)

    st.subheader("Table")
    sensitivity_display = sensitivity.copy()
    sensitivity_display.index = [f"{x * 100:.1f}%" for x in sensitivity.index]
    sensitivity_display.columns = [f"{x * 100:.1f}%" for x in sensitivity.columns]
    st.dataframe(sensitivity_display, use_container_width=True)

    st.subheader("Heatmap")
    fig_heatmap = px.imshow(
        sensitivity.values.astype(float),
        x=[f"{x * 100:.1f}%" for x in sensitivity.columns],
        y=[f"{y * 100:.1f}%" for y in sensitivity.index],
        labels=dict(x="WACC", y="Terminal Growth", color="Value per Share"),
        text_auto=".2f",
        aspect="auto",
        color_continuous_scale="RdYlGn",
    )
    fig_heatmap.update_xaxes(side="bottom")
    st.plotly_chart(fig_heatmap, use_container_width=True)


# =========================================================
# PAGE: Download
# =========================================================
def page_download():
    render_nav_bar()
    st.title("Download Valuation")
    st.caption("Export the full valuation to a multi-sheet Excel workbook with **live formulas** — edit any assumption in Excel and the whole valuation recalculates automatically.")

    inputs = build_inputs_from_state()
    if inputs.terminal_growth >= inputs.wacc:
        st.error("Terminal growth must be lower than WACC. Fix it on the Assumptions page first.")
        return
    if inputs.revenue_0 <= 0:
        st.error("Starting revenue is zero or unavailable, so the DCF cannot run.")
        return

    results = run_dcf(inputs)
    sensitivity = make_sensitivity_table(inputs)

    try:
        excel_bytes = build_excel_bytes(inputs, results, sensitivity)
        st.download_button(
            label="📥 Download valuation as Excel",
            data=excel_bytes,
            file_name=f"{inputs.ticker}_valuation.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.success(
            "File ready. The workbook contains 6 sheets: **Summary** (formulas), "
            "**Assumptions** (inputs you can edit), **Forecast** (formulas), "
            "**DCF Bridge** (formulas), **Sensitivity** (static), and **Read Me**. "
            "Edit any value on the Assumptions sheet and every downstream "
            "calculation updates automatically."
        )
    except ImportError as e:
        st.warning(f"Excel export is unavailable: {e}. Install with: `pip install openpyxl`")
    except Exception as e:
        st.error(f"Could not build Excel export: {e}")

    st.divider()
    with st.expander("How to replicate this valuation in Excel"):
        st.markdown(
            """
            Use the same assumptions and recreate these formulas row by row:

            1. **Revenue(t) = Revenue(t-1) × (1 + Growth Rate)**
            2. **EBIT = Revenue × EBIT Margin**
            3. **NOPAT = EBIT × (1 − Tax Rate)**
            4. **FCFF = NOPAT + D&A − Capex − Change in NWC**
            5. **PV of FCFF = FCFF / (1 + WACC)^t**
            6. **Terminal Value = FCFF(Year 5) × (1 + g) / (WACC − g)**
            7. **PV of Terminal Value = Terminal Value / (1 + WACC)^5**
            8. **Enterprise Value = Sum of PV of FCFF + PV of Terminal Value**
            9. **Equity Value = Enterprise Value − Net Debt**
            10. **Intrinsic Value per Share = Equity Value / Shares Outstanding**
            """
        )


# =========================================================
# Page dispatcher
# =========================================================
PAGE_MAP = {
    "Home": page_home,
    "Company Data": page_company_data,
    "Assumptions": page_assumptions,
    "Valuation": page_valuation,
    "Step-by-Step Breakdown": page_breakdown,
    "Sensitivity": page_sensitivity,
    "Download": page_download,
}

current_page = st.session_state.get("page", "Home")
PAGE_MAP.get(current_page, page_home)()